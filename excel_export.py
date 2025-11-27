# excel_export.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def generate_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan PO"

    # Write headers & rows using dataframe_to_rows
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        ws.append(row)
    # Style header
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Apply number format to total columns if they exist
    col_idx = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column+1)}
    for name in ["total_tagihan", "total_bayar", "sisa"]:
        if name in col_idx:
            c = col_idx[name]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = '#,##0'

    # Save to bytes
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
